import os
from openpyxl import load_workbook
import gurobipy as gp
from gurobipy import GRB
from itertools import combinations

EXCEL_PATH = "dados.xlsx"

def carregar_dados():
    dados = {
        "veiculos": {}, "faculdades": {}, "distancias": {},
        "lista_veiculos": [], "lista_faculdades": []
    }
    try:
        if not os.path.exists(EXCEL_PATH):
            raise FileNotFoundError(EXCEL_PATH)
        wb = load_workbook(EXCEL_PATH, data_only=True)

        ws_v = wb['Veiculos']
        headers_v = [c.value for c in ws_v[1]]
        idx = {h: i for i, h in enumerate(headers_v)}
        for row in ws_v.iter_rows(min_row=2, values_only=True):
            if not row or not row[idx['nome']]:
                continue
            nome = row[idx['nome']]
            dados['veiculos'][nome] = {
                'capacidade': int(row[idx['capacidade']]),
                'eficiencia_urbana': float(row[idx['eficiencia_urbana']]),
                'eficiencia_estrada': float(row[idx['eficiencia_estrada']]),
                'distancia_estrada': float(row[idx['distancia_estrada']])
            }
            dados['lista_veiculos'].append(nome)

        ws_f = wb['Faculdades']
        headers_f = [c.value for c in ws_f[1]]
        idxf = {h: i for i, h in enumerate(headers_f)}
        for row in ws_f.iter_rows(min_row=2, values_only=True):
            if not row or not row[idxf['nome']]:
                continue
            nome = row[idxf['nome']]
            dados['faculdades'][nome] = {'demanda': int(row[idxf['demanda']])}
            dados['lista_faculdades'].append(nome)

        ws_d = wb['Distancias']
        headers_d = [c.value for c in ws_d[1]]
        idxd = {h: i for i, h in enumerate(headers_d)}
        for row in ws_d.iter_rows(min_row=2, values_only=True):
            if not row or not row[idxd['origem']]:
                continue
            origem = row[idxd['origem']]
            destino = row[idxd['destino']]
            dist = float(row[idxd['distancia']])
            dados['distancias'][(origem, destino)] = dist

    except FileNotFoundError as e:
        print(f"Erro: Arquivo não encontrado - {e.filename}.")
        return None
    except (ValueError, KeyError) as e:
        print(f"Erro ao processar os dados dos arquivos: {e}")
        return None

    return dados


def preparar_modelo_dados(dados):
    origem = 'Bicas'
    nos = [origem] + dados['lista_faculdades']

    custo_arco = {}
    for v in dados['lista_veiculos']:
        veiculo_info = dados['veiculos'][v]
        for i in nos:
            for j in nos:
                if i == j: continue

                dist_urbana = dados['distancias'].get((i, j), 0)
                dist_estrada = veiculo_info['distancia_estrada']
                ef_urbano = veiculo_info['eficiencia_urbana']
                ef_estrada = veiculo_info['eficiencia_estrada']

                if i == origem and j in dados['lista_faculdades']:
                    custo = (dist_estrada / ef_estrada) + (dist_urbana / ef_urbano)
                elif i in dados['lista_faculdades'] and j in dados['lista_faculdades']:
                    custo = dist_urbana / ef_urbano
                else:
                    custo = 0

                custo_arco[v, i, j] = custo

    return {
        "veiculos": dados['lista_veiculos'], "faculdades": dados['lista_faculdades'], "nos": nos, "origem": origem,
        "capacidade": {v: d['capacidade'] for v, d in dados['veiculos'].items()},
        "demanda": {f: d['demanda'] for f, d in dados['faculdades'].items()},
        "custo_arco": custo_arco
    }


def resolver_modelo(modelo_dados):

    if not modelo_dados:
        return

    m = gp.Model("TransporteEscolar")

    # --- VARIÁVEIS DE DECISÃO ---

    # x[v, f]: Quantos alunos da faculdade 'f' o veículo 'v' transporta. (Decisão de Alocação)
    x = m.addVars(modelo_dados["veiculos"], modelo_dados["faculdades"], vtype=GRB.INTEGER, name="x")

    # y[v, i, j]: Binária. 1 se o veículo 'v' viaja do nó 'i' para o 'j'; 0 caso contrário. (Decisão de Roteirização)
    y = m.addVars(modelo_dados["veiculos"], modelo_dados["nos"], modelo_dados["nos"], vtype=GRB.BINARY, name="y")

    # delta[v, f]: Binária. 1 se o veículo 'v' para na faculdade 'f'; 0 caso contrário. (Variável Auxiliar de Ativação)
    delta = m.addVars(modelo_dados["veiculos"], modelo_dados["faculdades"], vtype=GRB.BINARY, name="delta")

    # --- FUNÇÃO OBJETIVO ---
    m.setObjective(gp.quicksum(modelo_dados["custo_arco"][v, i, j] * y[v, i, j]
                               for v, i, j in modelo_dados["custo_arco"]), GRB.MINIMIZE)

    # --- RESTRIÇÕES ---

    # 1. Atendimento da Demanda: Todos os alunos devem ser transportados.
    m.addConstrs((x.sum('*', f) == modelo_dados["demanda"][f] for f in modelo_dados["faculdades"]), "Demanda")

    # 2. Capacidade dos Veículos: O total de alunos em um veículo não pode exceder sua capacidade.
    m.addConstrs((x.sum(v, '*') <= modelo_dados["capacidade"][v] for v in modelo_dados["veiculos"]), "Capacidade")

    # 3. Conexão Alocação-Parada (Big-M): Um veículo só pode transportar alunos (x > 0) se ele parar na faculdade (delta = 1).
    m.addConstrs((x[v, f] <= modelo_dados["demanda"][f] * delta[v, f] for v in modelo_dados["veiculos"] for f in
                  modelo_dados["faculdades"]), "Conexao_BigM")

    # 4. Lógica de Roteamento: Se um veículo para numa faculdade (delta=1), exatamente um arco deve chegar nela.
    m.addConstrs(
        (gp.quicksum(y[v, i, f] for i in modelo_dados["nos"] if i != f) == delta[v, f] for v in modelo_dados["veiculos"]
         for f in modelo_dados["faculdades"]), "Ativacao_Parada")

    # 5. Garante que um arco só pode SAIR de uma parada ativada (delta=1).
    m.addConstrs(
        (y[v, i, j] <= delta[v, i] for v in modelo_dados["veiculos"] for i in modelo_dados["faculdades"] for j in
         modelo_dados["nos"] if j != i), "Arco_Saida_Requer_DeltaI")

    # 6. Garante que um arco só pode CHEGAR em uma parada ativada (delta=1).
    m.addConstrs(
        (y[v, i, j] <= delta[v, j] for v in modelo_dados["veiculos"] for j in modelo_dados["faculdades"] for i in
         modelo_dados["nos"] if j != i), "Arco_Chegada_Requer_DeltaJ")

    # 7. Parada Requer Atendimento: Se um veículo para (delta=1), deve transportar ao menos 1 aluno. Evita paradas vazias.
    m.addConstrs((x[v, f] >= delta[v, f] for v in modelo_dados["veiculos"] for f in modelo_dados["faculdades"]),
                 "Parada_Requer_Atendimento_Min1")

    # 8. Saída da Origem: Cada veículo pode sair da garagem (Bicas) no máximo uma vez.
    m.addConstrs((y.sum(v, modelo_dados["origem"], '*') <= 1 for v in modelo_dados["veiculos"]), "Saida_Origem")

    # 9. Se um veículo é USADO (soma de deltas > 0), ele OBRIGATORIAMENTE TEM que sair da origem.
    F = len(modelo_dados["faculdades"]) or 1
    m.addConstrs((y.sum(v, modelo_dados["origem"], '*') >= (1.0 / F) * gp.quicksum(
        delta[v, f] for f in modelo_dados["faculdades"]) for v in modelo_dados["veiculos"]), "Ativa_Origem_Se_Usado")

    # 10. Rota Aberta (Sem Retorno): Nenhum veículo pode voltar para a garagem.
    m.addConstrs((y.sum(v, '*', modelo_dados["origem"]) == 0 for v in modelo_dados["veiculos"]), "Sem_Retorno_Origem")

    # 11. Fluxo Contínuo: De cada faculdade, só pode sair um caminho. Evita que a rota se divida.
    m.addConstrs((gp.quicksum(y[v, i, j] for j in modelo_dados["nos"] if j != i) <= 1
                  for v in modelo_dados["veiculos"] for i in modelo_dados["faculdades"]), "Saida_Faculdade_AtMost1")

    # 12. Eliminação de Sub-rotas (DFJ): Impede que um veículo crie rotas circulares apenas entre faculdades, sem nunca partir da origem.
    for v in modelo_dados["veiculos"]:
        for tamanho_subconjunto in range(2, len(modelo_dados["faculdades"])):
            for subconjunto in combinations(modelo_dados["faculdades"], tamanho_subconjunto):
                m.addConstr(
                    gp.quicksum(y[v, i, j] for i in subconjunto for j in subconjunto if i != j) <= len(subconjunto) - 1)

    print("Iniciando a otimização do modelo...")
    m.optimize()

    exportar_resultados(m, modelo_dados, x, y)


def exportar_resultados(m, modelo_dados, x, y):
    try:
        x_vals = m.getAttr('X', x)
        y_vals = m.getAttr('X', y)
    except gp.GurobiError:
        x_vals, y_vals = {}, {}

    status_txt = {
        GRB.OPTIMAL: "ÓTIMO", GRB.INFEASIBLE: "INVIÁVEL", GRB.UNBOUNDED: "ILIMITADO",
    }.get(m.Status, f"STATUS={m.Status}")

    with open("resultados.txt", "w", encoding="utf-8") as f:
        f.write("==== RELATÓRIO DE RESULTADOS (Gurobi) ====\n\n")
        if m.Status == GRB.OPTIMAL:
            f.write(f"Status: {status_txt}\n")
            f.write(f"Custo Total de Combustível (litros): {m.ObjVal:.4f}\n")
        else:
            f.write(f"Status: {status_txt} — análise parcial/diagnóstico\n")
        f.write(f"Tempo de execução (s): {m.Runtime:.3f}\n\n")

        total_demanda = sum(modelo_dados["demanda"].values())
        f.write("-- Atendimento por Faculdade --\n")
        for fac in modelo_dados["faculdades"]:
            alocados = sum(x_vals.get((v, fac), 0) for v in modelo_dados["veiculos"]) if x_vals else 0
            f.write(f"  - {fac}: demanda {modelo_dados['demanda'][fac]}, alocados {int(round(alocados))}\n")
        f.write("\n")

        f.write("-- Detalhes por Veículo --\n")
        custo_total_reconstruido = 0.0
        for v in modelo_dados["veiculos"]:
            capacidade_v = modelo_dados['capacidade'][v]
            lotacao = sum(x_vals.get((v, fac), 0) for fac in modelo_dados["faculdades"]) if x_vals else 0
            if lotacao < 0.5 and m.Status == GRB.OPTIMAL:
                continue

            f.write(f"\nVeículo: {v}\n")
            f.write(
                f"  - Lotação: {int(round(lotacao))} / {capacidade_v} ({(lotacao / capacidade_v * 100 if capacidade_v else 0):.1f}%)\n")
            f.write("  - Alunos transportados: \n")
            for fac in modelo_dados["faculdades"]:
                qtd = x_vals.get((v, fac), 0) if x_vals else 0
                if qtd > 0.5:
                    f.write(f"      * {fac}: {int(round(qtd))} alunos\n")

            arcos_usados = []
            for i in modelo_dados["nos"]:
                for j in modelo_dados["nos"]:
                    if y_vals.get((v, i, j), 0) > 0.5:
                        arcos_usados.append((i, j))

            if not arcos_usados:
                f.write("  - Rota: não utilizada\n")
                continue

            succ = {i: j for i, j in arcos_usados}
            percurso = [modelo_dados["origem"]]
            atual = modelo_dados["origem"]
            while atual in succ:
                proximo = succ[atual]
                percurso.append(proximo)
                atual = proximo

            f.write("  - Rota: " + " -> ".join(percurso) + "\n")

            custo_v = 0.0
            f.write("  - Custos por arco (litros):\n")
            for i, j in arcos_usados:
                c = modelo_dados["custo_arco"].get((v, i, j), 0.0)
                f.write(f"      * {i} -> {j}: {c:.4f}\n")
                custo_v += c
            f.write(f"  - Custo total do veículo (litros): {custo_v:.4f}\n")
            custo_total_reconstruido += custo_v

        if m.Status == GRB.OPTIMAL:
            f.write("\nValidação: \n")
            f.write(f"  Soma dos custos por veículo: {custo_total_reconstruido:.4f} litros\n")
            f.write(f"  Objetivo do modelo:         {m.ObjVal:.4f} litros\n")
            f.write("  Diferença numérica pequena é esperada por arredondamentos.\n")

    print("Resultados salvos em 'resultados.txt'.")


if __name__ == "__main__":
    dados_lidos = carregar_dados()
    if dados_lidos:
        dados_para_modelo = preparar_modelo_dados(dados_lidos)
        resolver_modelo(dados_para_modelo)